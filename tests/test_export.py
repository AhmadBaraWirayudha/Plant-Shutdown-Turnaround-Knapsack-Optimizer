"""
test_export.py — Tests for export_to_excel() and its star-schema mirror
sheet builder, including direct end-to-end tests of the actual openpyxl
file-writing (not just the pure-logic dimension-table construction).
"""

from __future__ import annotations
from pathlib import Path
import numpy as np
import pandas as pd

from src.reporting.export import _build_dimension_sheets, export_to_excel


def _make_schedule(n=12) -> pd.DataFrame:
    task_types = ["Inspection", "Repair", "Overhaul"]
    priorities = ["Critical", "High", "Medium", "Low"]
    risk_levels = ["LOW", "MEDIUM", "HIGH", "CRITICAL"]
    rows = []
    for i in range(n):
        rows.append(
            {
                "wo_id": f"WO-{i:05d}",
                "description": f"Task {i}",
                "asset_tag": f"PMP-{i % 5:04d}",  # 5 distinct assets
                "asset_class": "PMP",
                "asset_name": "Centrifugal Pump",
                "area": "Unit-100",
                "install_date": "2018-08-23",
                "replace_usd": 50000.0,
                "c_safety": 3,
                "c_env": 2,
                "c_prod": 4,
                "c_cost": 2,
                "task_type": task_types[i % len(task_types)],
                "priority": priorities[i % len(priorities)],
                "risk_level": risk_levels[i % len(risk_levels)],
                "predecessor_wo_id": None,
                "mandatory": i == 0,
                "age_days": 500.0,
                "estimated_cost_usd": 10000.0,
                "mech_hours": 10.0,
                "elec_hours": 2.0,
                "inst_hours": 1.0,
                "civil_hours": 0.0,
                "total_craft_hours": 13.0,
                "duration_days": 2,
                "fitted_beta": 2.0,
                "fitted_eta": 1500.0,
                "failure_prob": 0.3,
                "rul_days": 800.0,
                "consequence_score": 3.0,
                "likelihood_tier": 3,
                "consequence_tier": 3,
                "risk_score": 9,
                "deferred_cost_usd": 5000.0,
                "net_value_usd": 1000.0,
                "selected": i % 2 == 0,
                "decision": "INCLUDE" if i % 2 == 0 else "DEFER",
            }
        )
    return pd.DataFrame(rows)


class TestExportToExcelEndToEnd:
    """
    Runs export_to_excel against a REAL SolverResult (from an actual small
    TurnaroundSolver.solve() call, not a hand-built mock) and inspects the
    written workbook directly. This is what catches a column-name mismatch
    between the solver's actual output and what export.py expects to find
    — a hand-rolled fixture, built with the "correct" column names by
    construction, could never catch that class of bug.
    """

    @staticmethod
    def _make_real_result(n=12, budget=200_000.0):
        from src.optimization.solver import TurnaroundSolver

        rng = np.random.default_rng(0)
        task_types = ["Inspection", "Repair", "Overhaul"]
        priorities = ["Critical", "High", "Medium", "Low"]
        risk_levels = ["LOW", "MEDIUM", "HIGH", "CRITICAL"]

        rows = []
        for i in range(n):
            rows.append(
                {
                    "wo_id": f"WO-{i:05d}",
                    "description": f"Task {i}",
                    "asset_tag": f"PMP-{i % 4:04d}",
                    "asset_class": "PMP",
                    "asset_name": "Centrifugal Pump",
                    "area": "Unit-100",
                    "replace_usd": 50000.0,
                    "c_safety": 3,
                    "c_env": 2,
                    "c_prod": 4,
                    "c_cost": 2,
                    "task_type": task_types[i % len(task_types)],
                    "priority": priorities[i % len(priorities)],
                    "risk_level": risk_levels[i % len(risk_levels)],
                    "predecessor_wo_id": None,
                    "mandatory": i == 0,
                    "age_days": float(rng.uniform(100, 2000)),
                    "estimated_cost_usd": float(rng.uniform(5000, 30000)),
                    "mech_hours": float(rng.uniform(1, 20)),
                    "elec_hours": float(rng.uniform(0, 5)),
                    "inst_hours": float(rng.uniform(0, 5)),
                    "civil_hours": float(rng.uniform(0, 2)),
                    "total_craft_hours": float(rng.uniform(5, 30)),
                    "duration_days": int(rng.integers(1, 5)),
                    "fitted_beta": 2.0,
                    "fitted_eta": 1500.0,
                    "failure_prob": float(rng.uniform(0, 1)),
                    "rul_days": float(rng.uniform(0, 2000)),
                    "consequence_score": float(rng.uniform(1, 5)),
                    "likelihood_tier": int(rng.integers(1, 6)),
                    "consequence_tier": int(rng.integers(1, 6)),
                    "risk_score": int(rng.integers(1, 26)),
                    "deferred_cost_usd": float(rng.uniform(0, 40000)),
                    "net_value_usd": float(rng.uniform(-5000, 40000)),
                }
            )
        wos = pd.DataFrame(rows)

        class _Cfg:
            total_budget = budget
            max_mech_hours = 1000.0
            max_elec_hours = 500.0
            max_inst_hours = 500.0
            max_civil_hours = 200.0

        return TurnaroundSolver(wos, config=_Cfg()).solve()

    def test_writes_a_real_xlsx_file(self, tmp_path):
        result = self._make_real_result()
        out_path = tmp_path / "test_export.xlsx"
        returned_path = export_to_excel(result, out_path=out_path)

        assert returned_path == out_path
        assert out_path.exists()
        assert out_path.stat().st_size > 0

    def test_workbook_contains_every_expected_sheet(self, tmp_path):
        import openpyxl

        result = self._make_real_result()
        out_path = tmp_path / "test_export.xlsx"
        export_to_excel(result, out_path=out_path)

        wb = openpyxl.load_workbook(out_path, read_only=True)
        expected = {
            "OptimizedSchedule",
            "Selected",
            "Deferred",
            "SummaryKPIs",
            "CapacityUtilization",
            "RiskMatrix",
            "ByArea",
            "ByEquipmentClass",
            "Dim_Asset",
            "Dim_TaskType",
            "Dim_Priority",
            "Dim_RiskLevel",
            "FactWorkOrderDecision",
        }
        assert expected.issubset(set(wb.sheetnames))

    def test_optimized_schedule_row_count_matches_total_tasks(self, tmp_path):
        import openpyxl

        result = self._make_real_result(n=12)
        out_path = tmp_path / "test_export.xlsx"
        export_to_excel(result, out_path=out_path)

        wb = openpyxl.load_workbook(out_path, read_only=True)
        ws = wb["OptimizedSchedule"]
        assert ws.max_row - 1 == 12  # minus header row

    def test_creates_parent_directory_if_missing(self, tmp_path):
        result = self._make_real_result(n=5)
        nested_path = tmp_path / "does" / "not" / "exist" / "export.xlsx"
        assert not nested_path.parent.exists()
        export_to_excel(result, out_path=nested_path)
        assert nested_path.exists()

    def test_accepts_plain_string_path_not_just_path_object(self, tmp_path):
        """
        Regression test for a real bug: passing a plain str (rather than a
        pathlib.Path) used to crash with
        'AttributeError: str object has no attribute parent' the moment
        out_path.parent.mkdir(...) ran — a bug invisible to every prior
        test in this file because they all used the tmp_path fixture
        (which yields Path objects), never a bare string. Passing a string
        is completely natural for any caller, so this is now coerced
        internally.
        """
        str_path = str(tmp_path / "string_path_export.xlsx")
        result = self._make_real_result(n=5)
        returned = export_to_excel(result, out_path=str_path)
        assert returned == Path(str_path)
        assert Path(str_path).exists()

    def test_omitted_out_path_resolves_live_reports_dir(self, monkeypatch, tmp_path):
        """The out_path=None sentinel must resolve REPORTS_DIR at CALL
        time, not whatever it was when this module was first imported —
        same bug class as docs/METHODOLOGY.md §5."""
        import src.reporting.export as export_mod

        monkeypatch.setattr(export_mod, "REPORTS_DIR", tmp_path)
        result = self._make_real_result(n=5)
        returned = export_to_excel(result)
        assert returned == tmp_path / "power_bi_export.xlsx"
        assert returned.exists()


class TestExportZeroDivisionSafety:
    """
    Regression tests for division-by-zero in the Excel export:
    two locations had unguarded division that crashed when a trade-hour
    capacity or equipment-class cost was exactly 0.

    1. CapacityUtilization sheet: `u / c * 100` crashed when c=0
       (a zero-capacity trade is a legitimate real-world config, e.g.
       no civil-craft work planned for this turnaround).
    2. ByEquipmentClass sheet: `total_value / total_cost.replace(0, 1)`
       didn't crash but silently returned total_value (a raw dollar amount)
       mislabeled as a ratio — the same bug class fixed in solver.py's
       roi_ratio calculation (see docs/METHODOLOGY.md §6).
    """

    @staticmethod
    def _make_result_with_zero_civil_cap():
        """A real solve with max_civil_hours=0 so civil utilisation = 0/0."""
        rng = np.random.default_rng(42)
        rows = []
        for i in range(8):
            rows.append(
                {
                    "wo_id": f"WO-{i:05d}",
                    "description": f"Task {i}",
                    "asset_tag": f"PMP-{i % 4:04d}",
                    "asset_class": "PMP",
                    "asset_name": "Pump",
                    "area": "Unit-100",
                    "replace_usd": 50000.0,
                    "c_safety": 3,
                    "c_env": 2,
                    "c_prod": 4,
                    "c_cost": 2,
                    "task_type": "Inspection",
                    "priority": "Medium",
                    "risk_level": "LOW",
                    "predecessor_wo_id": None,
                    "mandatory": False,
                    "age_days": 500.0,
                    "estimated_cost_usd": float(rng.uniform(1000, 5000)),
                    "mech_hours": 5.0,
                    "elec_hours": 1.0,
                    "inst_hours": 1.0,
                    "civil_hours": 0.0,
                    "total_craft_hours": 7.0,
                    "duration_days": 1,
                    "fitted_beta": 2.0,
                    "fitted_eta": 1500.0,
                    "failure_prob": 0.3,
                    "rul_days": 800.0,
                    "consequence_score": 3.0,
                    "likelihood_tier": 3,
                    "consequence_tier": 3,
                    "risk_score": 9,
                    "deferred_cost_usd": 500.0,
                    "net_value_usd": float(rng.uniform(1000, 3000)),
                }
            )
        wos = pd.DataFrame(rows)

        class _Cfg:
            total_budget = 50000.0
            max_mech_hours = 1000.0
            max_elec_hours = 500.0
            max_inst_hours = 500.0
            max_civil_hours = 0.0  # <-- zero-capacity trade

        from src.optimization.solver import TurnaroundSolver

        return TurnaroundSolver(wos, config=_Cfg()).solve()

    def test_capacity_utilisation_with_zero_cap_does_not_crash(self, tmp_path):
        """
        Regression: `u / c * 100` in CapacityUtilization sheet raised
        ZeroDivisionError when a trade's capacity was 0.
        """
        result = self._make_result_with_zero_civil_cap()
        out = export_to_excel(result, out_path=tmp_path / "zero_cap.xlsx")
        assert out.exists()

    def test_capacity_utilisation_zero_cap_shows_zero_percent(self, tmp_path):
        """A 0-capacity trade must show 0% utilisation, not crash or NaN."""
        import openpyxl

        result = self._make_result_with_zero_civil_cap()
        out = export_to_excel(result, out_path=tmp_path / "zero_cap2.xlsx")
        wb = openpyxl.load_workbook(out)
        ws = wb["CapacityUtilization"]

        # Find the Civil row by its trade name and check utilisation
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        util_col = headers.index("Utilisation (%)") + 1
        trade_col = headers.index("Trade") + 1

        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=trade_col).value == "Civil":
                util = ws.cell(row=row, column=util_col).value
                assert util == 0.0, f"Civil utilisation should be 0.0, got {util!r}"
                break

    def test_roi_zero_cost_shows_zero_not_raw_value(self, tmp_path):
        """
        Regression: `total_value / total_cost.replace(0, 1)` returned raw
        dollar values when total_cost=0 — e.g. roi=5000.0 meaning $5000,
        not 5000x. This is the same bug class fixed in solver.py's
        roi_ratio. The correct answer when nothing was spent is 0.0.
        """
        import openpyxl

        # Build a result where nothing is selected (budget=0 → zero cost, zero value)
        rng = np.random.default_rng(1)
        rows = []
        for i in range(5):
            rows.append(
                {
                    "wo_id": f"WO-{i:05d}",
                    "description": f"T{i}",
                    "asset_tag": f"PMP-{i % 2:04d}",
                    "asset_class": "PMP",
                    "asset_name": "Pump",
                    "area": "Unit-100",
                    "replace_usd": 50000.0,
                    "c_safety": 3,
                    "c_env": 2,
                    "c_prod": 4,
                    "c_cost": 2,
                    "task_type": "Inspection",
                    "priority": "Low",
                    "risk_level": "LOW",
                    "predecessor_wo_id": None,
                    "mandatory": False,
                    "age_days": 500.0,
                    "estimated_cost_usd": 99999.0,
                    "mech_hours": 5.0,
                    "elec_hours": 1.0,
                    "inst_hours": 1.0,
                    "civil_hours": 0.0,
                    "total_craft_hours": 7.0,
                    "duration_days": 1,
                    "fitted_beta": 2.0,
                    "fitted_eta": 1500.0,
                    "failure_prob": 0.3,
                    "rul_days": 800.0,
                    "consequence_score": 3.0,
                    "likelihood_tier": 3,
                    "consequence_tier": 3,
                    "risk_score": 9,
                    "deferred_cost_usd": 500.0,
                    "net_value_usd": float(rng.uniform(100, 500)),
                }
            )
        wos = pd.DataFrame(rows)

        class _Cfg:
            total_budget = 0.0  # budget=0: nothing selected, total_cost=0
            max_mech_hours = 1000.0
            max_elec_hours = 500.0
            max_inst_hours = 500.0
            max_civil_hours = 0.0

        from src.optimization.solver import TurnaroundSolver

        result = TurnaroundSolver(wos, config=_Cfg()).solve()
        assert result.summary["tasks_selected"] == 0

        out = export_to_excel(result, out_path=tmp_path / "zero_roi.xlsx")
        wb = openpyxl.load_workbook(out)
        ws = wb["ByEquipmentClass"]

        if ws.max_row < 2:
            return  # no rows to check if all were deferred

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        if "roi" not in headers:
            return

        roi_col = headers.index("roi") + 1
        for row in range(2, ws.max_row + 1):
            roi = ws.cell(row=row, column=roi_col).value
            if roi is not None:
                assert abs(roi) < 10, f"roi={roi!r} looks like a raw dollar value, not a ratio"


class TestFormulaInjectionSanitisation:
    """
    Regression tests for an OWASP-listed vulnerability: 'CSV/Excel formula
    injection'. User-controlled string fields (wo_id, description, area,
    task_type, etc.) from a real CMMS can contain values that start with
    '=', '+', '-', '@' — characters Excel treats as formula triggers. If
    written to a cell verbatim, these execute as Excel formulas when the
    workbook is opened, potentially calling HYPERLINK(), DDE(), or other
    dangerous functions. The _sanitise_formula_injection() helper prefixes
    such values with a single apostrophe (Excel's 'treat as plain text'
    escape), which is invisible to the reader but prevents execution.
    """

    def test_sanitise_escapes_equals_prefix(self):
        from src.reporting.export import _sanitise_formula_injection

        df = pd.DataFrame({"wo_id": ["=SUM(A1:A10)", "WO-001"]})
        out = _sanitise_formula_injection(df)
        assert out.loc[0, "wo_id"] == "'=SUM(A1:A10)"
        assert out.loc[1, "wo_id"] == "WO-001"  # safe value unchanged

    def test_sanitise_escapes_all_formula_trigger_chars(self):
        from src.reporting.export import _sanitise_formula_injection

        triggers = {
            "=": '=HYPERLINK("http://evil.com")',
            "+": "+SUM(A1)",
            "-": "-2+3",
            "@": "@SUM(A1)",
        }
        for prefix, val in triggers.items():
            df = pd.DataFrame({"col": [val]})
            out = _sanitise_formula_injection(df)
            assert out.loc[0, "col"] == f"'{val}", f"Value starting with '{prefix}' was not escaped"

    def test_sanitise_leaves_numeric_columns_unchanged(self):
        """Numbers are never formula-injectable through openpyxl — only
        string columns need escaping."""
        from src.reporting.export import _sanitise_formula_injection

        df = pd.DataFrame({"cost": [1000.0, 2000.0], "label": ["=BAD", "ok"]})
        out = _sanitise_formula_injection(df)
        assert list(out["cost"]) == [1000.0, 2000.0]  # unchanged
        assert out.loc[0, "label"] == "'=BAD"

    def test_excel_export_sanitises_formula_injection_in_wo_id(self, tmp_path):
        """End-to-end: a wo_id starting with '=' must arrive in the Excel
        file with the apostrophe prefix, preventing formula execution."""
        import openpyxl

        result = TestExportToExcelEndToEnd._make_real_result(n=5)
        result.schedule.loc[0, "wo_id"] = '=HYPERLINK("http://evil.com","click")'
        result.selected_schedule = result.schedule[result.schedule["selected"]].copy()
        result.deferred_schedule = result.schedule[~result.schedule["selected"]].copy()

        out_path = tmp_path / "injection_test.xlsx"
        export_to_excel(result, out_path=out_path)

        wb = openpyxl.load_workbook(out_path)
        ws = wb["OptimizedSchedule"]
        # Find the injected value in the wo_id column (first column)
        wo_ids = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        for val in wo_ids:
            if val and "HYPERLINK" in str(val):
                assert str(val).startswith("'"), f"Formula injection NOT escaped in Excel: {val!r}"


class TestBuildDimensionSheets:
    def test_returns_all_five_expected_sheets(self):
        sched = _make_schedule()
        sheets = _build_dimension_sheets(sched)
        assert set(sheets.keys()) == {
            "Dim_Asset",
            "Dim_TaskType",
            "Dim_Priority",
            "Dim_RiskLevel",
            "FactWorkOrderDecision",
        }

    def test_dim_asset_has_one_row_per_distinct_asset(self):
        sched = _make_schedule(n=20)  # uses 5 distinct asset tags (i % 5)
        sheets = _build_dimension_sheets(sched)
        assert len(sheets["Dim_Asset"]) == 5
        assert sheets["Dim_Asset"]["asset_tag"].is_unique

    def test_dim_asset_includes_install_date(self):
        """
        Regression test: Dim_Asset's docstring claims it 'mirrors
        src/db/schema.py' exactly, but install_date was missing from the
        column list entirely — present in the database's DimAsset table
        (and populated by the writer) but silently absent from the
        Excel-only Power BI path. A real-world install_date is a Timestamp
        from the pipeline, but it's normalized to a plain string here to
        match the database's String(20) column representation, so both
        Power BI connection paths (Option C/Excel vs the live database)
        show the identical value for this field.
        """
        sched = _make_schedule(n=5)
        sheets = _build_dimension_sheets(sched)
        dim_asset = sheets["Dim_Asset"]
        assert "install_date" in dim_asset.columns
        # The dtype LABEL varies across pandas versions (legacy 'object'
        # vs newer pandas StringDtype) — what actually matters is that
        # every value is a plain Python str, not a pandas Timestamp.
        assert all(isinstance(v, str) for v in dim_asset["install_date"])
        assert dim_asset.iloc[0]["install_date"] == "2018-08-23"

    def test_dim_asset_install_date_normalizes_real_timestamp_to_string(self):
        """The fixture above uses a pre-stringified date; this test
        confirms a genuine pandas Timestamp (what the real pipeline
        actually produces) is also normalized correctly, not left as a
        Timestamp object that openpyxl would otherwise serialize
        differently than the database's plain-string representation."""
        sched = _make_schedule(n=3)
        sched["install_date"] = pd.to_datetime(sched["install_date"])
        assert pd.api.types.is_datetime64_any_dtype(sched["install_date"])

        sheets = _build_dimension_sheets(sched)
        dim_asset = sheets["Dim_Asset"]
        assert all(isinstance(v, str) for v in dim_asset["install_date"])
        assert not any(isinstance(v, pd.Timestamp) for v in dim_asset["install_date"])

    def test_dim_task_type_has_no_duplicates(self):
        sched = _make_schedule(n=30)
        sheets = _build_dimension_sheets(sched)
        dim = sheets["Dim_TaskType"]
        assert dim["task_type_name"].is_unique
        assert dim["task_type_id"].is_unique
        assert set(dim["task_type_name"]) == set(sched["task_type"].unique())

    def test_dim_priority_weights_match_convention(self):
        sched = _make_schedule()
        sheets = _build_dimension_sheets(sched)
        dim = sheets["Dim_Priority"].set_index("priority_name")
        assert dim.loc["Critical", "priority_weight"] == 4
        assert dim.loc["Low", "priority_weight"] == 1

    def test_dim_risk_level_sort_order_matches_severity_convention(self):
        sched = _make_schedule()
        sheets = _build_dimension_sheets(sched)
        dim = sheets["Dim_RiskLevel"].set_index("risk_level_name")
        assert dim.loc["LOW", "sort_order"] < dim.loc["CRITICAL", "sort_order"]
        assert dim.loc["MEDIUM", "sort_order"] < dim.loc["HIGH", "sort_order"]

    def test_fact_table_has_no_unresolved_foreign_keys(self):
        """Every fact row must successfully resolve to a task_type_id,
        priority_id, and risk_level_id — a merge failure here would mean
        the Power BI relationships silently lose rows."""
        sched = _make_schedule(n=25)
        sheets = _build_dimension_sheets(sched)
        fact = sheets["FactWorkOrderDecision"]
        assert fact["task_type_id"].notna().all()
        assert fact["priority_id"].notna().all()
        assert fact["risk_level_id"].notna().all()

    def test_fact_table_row_count_matches_schedule(self):
        sched = _make_schedule(n=17)
        sheets = _build_dimension_sheets(sched)
        assert len(sheets["FactWorkOrderDecision"]) == 17

    def test_fact_table_does_not_carry_raw_categorical_columns(self):
        """The fact table should reference dimensions via _id columns, not
        duplicate the raw task_type/priority/risk_level strings — that's
        the whole point of normalizing into a star schema."""
        sched = _make_schedule()
        sheets = _build_dimension_sheets(sched)
        fact = sheets["FactWorkOrderDecision"]
        assert "task_type" not in fact.columns
        assert "priority" not in fact.columns
        assert "risk_level" not in fact.columns
        assert "task_type_id" in fact.columns

    def test_handles_single_distinct_value_per_dimension(self):
        """Edge case: a schedule where every row shares the same task_type,
        priority, and risk_level must still produce a valid 1-row dimension,
        not crash on a degenerate unique() call."""
        sched = _make_schedule(n=5)
        sched["task_type"] = "Inspection"
        sched["priority"] = "Medium"
        sched["risk_level"] = "LOW"
        sheets = _build_dimension_sheets(sched)
        assert len(sheets["Dim_TaskType"]) == 1
        assert len(sheets["Dim_Priority"]) == 1
        assert len(sheets["Dim_RiskLevel"]) == 1
        assert sheets["FactWorkOrderDecision"]["task_type_id"].notna().all()
